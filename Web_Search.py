#!/usr/bin/python 
import sys
import time
import urllib2
import logging
import datetime
import os
from prettytable import PrettyTable
import csv
from openpyxl import Workbook
 
import xlrd
import xlwt
from xlutils.copy import copy
import os.path 

class myClass:

	def __init__(self):
		pass

	def parseAddress(self, input):
		if input[:10] == "http://www":
			return input
		else:
			print "Error: Cannot retrieve URL, protocol must be HTTP"
			sys.exit(1)	
		
	def myMethod(self, text, website, wtime):
		# Open the log file 
		LOG_FILENAME = 'Log_file.log'
		headers = ["Input text", "start time","Found in", "Number of words Found", "Time of Execution"]
		t = PrettyTable(headers)

		wb=Workbook()
		ws = wb.active
	
		# Writing to File
		logFile = open(LOG_FILENAME, "a")# "a" (append) mode to avoid overwriting
		# for header in headers:
		# 	logFile.write('|\t\t\t' + header + '\t\t\t') # write the headers.
		# logFile.write('\n\n')
		start_time = time.time() # Current time
		ended = False
		print(start_time)
		print(time.time())
		while ended == False:
			for site in website:
				urlAddress = self.parseAddress(site)
				req = urllib2.urlopen(urlAddress).read()
				count = req.find(text)
				if count == -1:
					print("Can not found given text")
				else:  
					print(text)
					print(count)
					timeDelta = time.time() - start_time
					print timeDelta
					t.add_row([text, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), urlAddress,  str(count), str(timeDelta)])
				
					#with open('mytable.csv', 'ab') as MyTable:
					#	wr = csv.writer(MyTable, delimiter=';', quoting=csv.QUOTE_ALL)
					#	wr.writerow(headers)
					#	wr.writerow([text, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), urlAddress,  str(count), str(timeDelta)])


					#rb = xlrd.open_workbook('workbook.xlsx',formatting_info=True)
					#wb = copy(rb) 
					#ws = wb.get_sheet(0) 

					ws['A1'] = headers
					ws['A2'] = [text, atetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), urlAddress,  str(count), str(timeDelta)]
					ws.append(ws['A2'])
					wb.save('workbook.xlsx')
			
			
							
					# logFile.write(datetime.datetime.now().strftime('|\t\t' + "%Y-%m-%d %H:%M:%S" + '\t\t' ) + '|\t\t' +  urlAddress + '\t\t' ' ')
					# logFile.write('|\t\t' + (text) + ' = ' + str(count) + ' times '  + '\t\t'  + '|\t\t' + str(timeDelta) + '\t\t' '\n')
			

			ended = True
		logFile.write(str(t) + '\n')
